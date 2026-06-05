#!/usr/bin/env python3
"""
生成 lmbench 4 配置 × N=10 干净数据的综合 xlsx。
3 个 sheet:
  1. All metrics  —— 684 项全部指标的统计 + 跨配置 Δ%
  2. Highlights   —— 论文表对照的 30+ 项 + 关键 mmap/cache/bw
  3. Per-iter     —— 关键指标的 10 iter × 4 配置原始单值

参数：
  python3 build-xlsx.py [median|mean]
  median (默认)：中位数 + MAD%，输出 lmbench-N10-4config.xlsx
  mean       ：均值 + 相对标准差 (RSD%)，输出 lmbench-N10-4config-mean.xlsx
"""
import csv, statistics, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

STAT_MODE = sys.argv[1] if len(sys.argv) > 1 else 'median'
assert STAT_MODE in ('median', 'mean'), f'STAT_MODE 必须是 median 或 mean，给的是 {STAT_MODE}'
STAT_LABEL = '中位' if STAT_MODE == 'median' else '均值'
DISP_LABEL = 'MAD%' if STAT_MODE == 'median' else 'RSD%'  # MAD% = median abs deviation %; RSD% = relative std dev %

DAY3 = 'results/n90-day3'
CONFIGS = ['kvmoff', 'nvhe', 'vhe', 'pkvm']
# Note: pkvm 用 try2（同配置同方法的复跑数据），不用 try1
# try1 偶然碰到一次性 stochastic 加速状态（详见 standalone-memory-bench-validation.md §9），
# try2 完全没复现 → try2 是 pkvm 真实代表性数据
CSVS = {
    'kvmoff': f'{DAY3}/n90-kvmoff-noLSM-full-cpu0.csv',
    'nvhe':   f'{DAY3}/n90-nvhe-noLSM-full-cpu0.csv',
    'vhe':    f'{DAY3}/n90-vhe-noLSM-full-cpu0.csv',
    'pkvm':   f'{DAY3}/n90-pkvm-noLSM-try2-cpu0.csv',
}

def load(p):
    g = defaultdict(list)
    for r in csv.DictReader(open(p)):
        try:
            g[(r['bench'], r['variant'], r['unit'])].append((int(r['iter']), float(r['value'])))
        except: pass
    return g

raw = {c: load(p) for c, p in CSVS.items()}

def med_mad(xs):
    if not xs: return None, None
    vals = sorted(v for _, v in xs)
    m = statistics.median(vals)
    if not m: return m, None
    mad = 100 * statistics.median([abs(x-m) for x in vals]) / m
    return m, mad

def mean_stddev(xs):
    """返回 (均值, 相对标准差 %)。"""
    if not xs: return None, None
    vals = [v for _, v in xs]
    if len(vals) < 2: return vals[0] if vals else None, 0.0
    m = statistics.mean(vals)
    if not m: return m, None
    sd = statistics.stdev(vals)
    return m, 100 * sd / m

def stat(xs):
    """根据 STAT_MODE 选用 med_mad 或 mean_stddev。"""
    return med_mad(xs) if STAT_MODE == 'median' else mean_stddev(xs)

def iter_value(xs, iter_n):
    for i, v in xs:
        if i == iter_n: return v
    return None

# ============ styling ============
HEAD_FILL = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
HEAD_FONT = Font(color='FFFFFF', bold=True)
GROUP_FILL = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
GROUP_FONT = Font(bold=True)
THIN = Border(left=Side(style='thin', color='CCCCCC'),
              right=Side(style='thin', color='CCCCCC'),
              top=Side(style='thin', color='CCCCCC'),
              bottom=Side(style='thin', color='CCCCCC'))

# 颜色编码：方向 × 强度
# 方向：latency / smaller-is-better → 负 = 绿；正 = 红
#       bandwidth / larger-is-better → 正 = 绿；负 = 红
# 强度：|Δ| > 10% 深 + 粗体；5-10% 中；2-5% 浅；≤2% 黑（噪声）
GREEN_STRONG = Font(color='006100', bold=True)
GREEN_MED    = Font(color='2E7D32')
GREEN_LIGHT  = Font(color='66BB6A')
RED_STRONG   = Font(color='9C0006', bold=True)
RED_MED      = Font(color='C00000')
RED_LIGHT    = Font(color='E26B0A')
NEUTRAL      = Font(color='000000')

def is_lower_better(unit):
    """True if smaller value = better (latency)."""
    if unit in ('us', 'ns', 'seconds', 'sec'):
        return True
    # MB/s, KB/s, pages, count: higher = better
    return False

def delta_font(delta_pct, unit):
    """Pick font color based on Δ% magnitude and whether unit is lower-is-better."""
    if delta_pct is None: return NEUTRAL
    mag = abs(delta_pct)
    if mag <= 2:
        return NEUTRAL
    # 方向：是否 "好" 信号
    if is_lower_better(unit):
        is_good = delta_pct < 0  # 越小越好 → 负 Δ% = 更快 = 绿
    else:
        is_good = delta_pct > 0  # 越大越好 → 正 Δ% = 更高 = 绿
    if mag > 10:
        return GREEN_STRONG if is_good else RED_STRONG
    elif mag > 5:
        return GREEN_MED if is_good else RED_MED
    else:
        return GREEN_LIGHT if is_good else RED_LIGHT

def set_header(ws, row_idx, labels):
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row=row_idx, column=ci, value=lbl)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN
    ws.row_dimensions[row_idx].height = 32

def freeze_and_filter(ws, freeze_cell, last_col, last_row):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = f'A1:{get_column_letter(last_col)}{last_row}'

def auto_width(ws, cols, default=12):
    for ci, w in cols.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

# ============ Sheet 1: All Metrics ============
def build_all_metrics(wb):
    ws = wb.create_sheet('All metrics')
    headers = [
        'bench', 'variant', 'unit',
        f'kvmoff {STAT_LABEL}', f'kvmoff {DISP_LABEL}',
        f'nvhe {STAT_LABEL}', f'nvhe {DISP_LABEL}',
        f'vhe {STAT_LABEL}', f'vhe {DISP_LABEL}',
        f'pkvm {STAT_LABEL}', f'pkvm {DISP_LABEL}',
        'Δ% nvhe vs kvmoff',
        'Δ% vhe vs kvmoff',
        'Δ% pkvm vs kvmoff',
        'Δ% pkvm vs vhe',
        'Δ% pkvm vs nvhe',
    ]
    set_header(ws, 1, headers)

    all_keys = sorted(set().union(*[set(raw[c].keys()) for c in CONFIGS]))
    for ri, k in enumerate(all_keys, 2):
        bench, variant, unit = k
        km, kmad = stat(raw['kvmoff'].get(k, []))
        nm, nmad = stat(raw['nvhe'].get(k, []))
        vm, vmad = stat(raw['vhe'].get(k, []))
        pm, pmad = stat(raw['pkvm'].get(k, []))
        def d(a, b): return 100*(a-b)/b if (a is not None and b not in (None, 0)) else None
        row = [bench, variant, unit, km, kmad, nm, nmad, vm, vmad, pm, pmad,
               d(nm, km), d(vm, km), d(pm, km), d(pm, vm), d(pm, nm)]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN
            if ci >= 4 and ci <= 11 and val is not None:
                c.number_format = '0.0000'
            if ci >= 12 and val is not None:
                c.number_format = '+0.00;-0.00'
                c.font = delta_font(val, unit)

    freeze_and_filter(ws, 'D2', len(headers), len(all_keys)+1)
    auto_width(ws, {
        1: 22, 2: 28, 3: 8,
        4: 14, 5: 11, 6: 14, 7: 11, 8: 14, 9: 11, 10: 14, 11: 11,
        12: 17, 13: 17, 14: 17, 15: 17, 16: 17,
    })

# ============ Sheet 2: Highlights ============
PAPER_TABLE = [
    ('系统调用与进程', [
        ('null call (空系统调用)',         'lat_syscall', 'null', 'us'),
        ('open/close (文件打开关闭)',       'lat_syscall', 'open', 'us'),
        ('fork (进程创建)',                 'lat_proc', 'fork', 'us'),
        ('exec (进程执行)',                 'lat_proc', 'exec', 'us'),
        ('sh (Shell 脚本执行)',             'lat_proc', 'shell', 'us'),
    ]),
    ('处理器运算', [
        ('intgr add',                       'lat_ops', 'integer_add', 'ns'),
        ('intgr div',                       'lat_ops', 'integer_div', 'ns'),
        ('double add',                      'lat_ops', 'double_add', 'ns'),
        ('double div',                      'lat_ops', 'double_div', 'ns'),
    ]),
    ('上下文切换', [
        ('2p/0K ctxsw',                     'lat_ctx', 'sz0k_p2', 'us'),
        ('2p/16K ctxsw',                    'lat_ctx', 'sz16k_p2', 'us'),
        ('8p/16K ctxsw',                    'lat_ctx', 'sz16k_p8', 'us'),
        ('16p/16K ctxsw',                   'lat_ctx', 'sz16k_p16', 'us'),
    ]),
    ('本地通信延迟', [
        ('Pipe (管道)',                     'lat_pipe', 'rt', 'us'),
        ('AF UNIX (域套接字)',              'lat_unix', 'rt', 'us'),
        ('UDP',                             'lat_udp', 'lhost', 'us'),
        ('TCP',                             'lat_tcp', 'lhost', 'us'),
        ('TCP conn (TCP 连接)',             'lat_connect', 'lhost', 'us'),
    ]),
    ('文件系统（ops/s，越大越好）', [
        ('0K File Create (ops/s)',          'lat_fs', 'sz0K_create', 'ops/s'),
        ('10K File Create (ops/s)',         'lat_fs', 'sz10K_create', 'ops/s'),
        ('0K File Unlink (ops/s)',          'lat_fs', 'sz0K_unlink', 'ops/s'),
        ('10K File Unlink (ops/s)',         'lat_fs', 'sz10K_unlink', 'ops/s'),
    ]),
    ('内存延迟', [
        ('Prot Fault (保护错误)',           'lat_sig', 'prot', 'us'),
        ('Page Fault (缺页中断)',           'lat_pagefault', 'minor', 'us'),
        ('L1 Cache (~64KB, stride256)',     'lat_mem_rd_load', 'stride256_sz0.06250MB', 'ns'),
        ('L2 Cache (~512KB, stride256)',    'lat_mem_rd_load', 'stride256_sz0.50000MB', 'ns'),
        ('LLC (~1MB, stride256)',           'lat_mem_rd_load', 'stride256_sz1.00000MB', 'ns'),
        ('LLC 边界 (~8MB)',                  'lat_mem_rd_load', 'stride256_sz8.00000MB', 'ns'),
        ('Main Memory (顺序, 64MB)',        'lat_mem_rd_load', 'stride256_sz64.00000MB', 'ns'),
        ('Random Memory (随机, 64MB stride16)', 'lat_mem_rd_rand', 'stride16_sz64.00000MB', 'ns'),
        ('Random Memory (随机, 8MB stride16)',  'lat_mem_rd_rand', 'stride16_sz8.00000MB', 'ns'),
        ('Random Memory (随机, 4MB stride16)',  'lat_mem_rd_rand', 'stride16_sz4.00000MB', 'ns'),
    ]),
    ('mmap 延迟 (pKVM stage-2 信号)', [
        ('mmap 0.5 MB',                     'lat_mmap', 'sz0.524288MB', 'us'),
        ('mmap 1 MB',                       'lat_mmap', 'sz1.048576MB', 'us'),
        ('mmap 16 MB',                      'lat_mmap', 'sz16.777216MB', 'us'),
        ('mmap 67 MB',                      'lat_mmap', 'sz67.108864MB', 'us'),
    ]),
    ('内存带宽 (MB/s)', [
        ('Pipe bw',                         'bw_pipe', 'bw', 'MB/s'),
        ('TCP bw (10MB msg)',               'bw_tcp', 'msg10.000000MB', 'MB/s'),
        ('Mem read (DRAM 67MB)',            'bw_mem_rd', 'sz67.11', 'MB/s'),
        ('Mem write (DRAM 67MB)',           'bw_mem_wr', 'sz67.11', 'MB/s'),
        ('Mem bzero (DRAM 67MB)',           'bw_mem_bzero', 'sz67.11', 'MB/s'),
        # peak sizes：每个 bench 自己在 cache hierarchy 上的实测峰值位置
        ('Mem rd peak (cache resident)',    'bw_mem_rd', 'sz0.065536', 'MB/s'),
        ('Mem wr peak (cache resident)',    'bw_mem_wr', 'sz0.032768', 'MB/s'),
        ('Mem bzero peak (cache resident)', 'bw_mem_bzero', 'sz0.131072', 'MB/s'),
    ]),
    ('TLB', [
        ('tlb effective (pages)',           'tlb', 'effective', 'pages'),
    ]),
]

def build_highlights(wb):
    ws = wb.create_sheet('Highlights', 0)  # first sheet
    headers = ['维度', '具体测试项', 'unit',
               'kvmoff', 'nvhe', 'vhe', 'pkvm',
               'Δ% nvhe/kvmoff', 'Δ% vhe/kvmoff', 'Δ% pkvm/kvmoff',
               'Δ% pkvm/vhe', 'Δ% pkvm/nvhe']
    set_header(ws, 1, headers)
    ri = 2
    for section, items in PAPER_TABLE:
        # section header row
        c = ws.cell(row=ri, column=1, value=section)
        c.fill = GROUP_FILL; c.font = GROUP_FONT
        for ci in range(2, len(headers)+1):
            cc = ws.cell(row=ri, column=ci); cc.fill = GROUP_FILL
        ri += 1
        for label, bench, variant, unit in items:
            k = (bench, variant, unit)
            km, _ = stat(raw['kvmoff'].get(k, []))
            nm, _ = stat(raw['nvhe'].get(k, []))
            vm, _ = stat(raw['vhe'].get(k, []))
            pm, _ = stat(raw['pkvm'].get(k, []))
            def d(a, b): return 100*(a-b)/b if (a is not None and b not in (None, 0)) else None
            row = ['', label, unit, km, nm, vm, pm,
                   d(nm, km), d(vm, km), d(pm, km),
                   d(pm, vm), d(pm, nm)]
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = THIN
                if ci >= 4 and ci <= 7 and val is not None:
                    c.number_format = '0.0000'
                if ci >= 8 and val is not None:
                    c.number_format = '+0.00;-0.00'
                    c.font = delta_font(val, unit)
            ri += 1
    freeze_and_filter(ws, 'D2', len(headers), ri-1)
    auto_width(ws, {
        1: 24, 2: 36, 3: 8,
        4: 12, 5: 12, 6: 12, 7: 12,
        8: 17, 9: 17, 10: 17, 11: 17, 12: 17,
    })

# ============ Sheet 3: Per-iter raw ============
def build_per_iter(wb):
    ws = wb.create_sheet('Per-iter raw')
    # only key metrics
    flat = [(label, b, v, u) for sec, items in PAPER_TABLE for label, b, v, u in items]
    headers = ['指标', 'unit', 'config'] + [f'iter {i}' for i in range(1, 11)] + [STAT_LABEL, DISP_LABEL]
    set_header(ws, 1, headers)
    ri = 2
    for label, b, v, u in flat:
        for cfg in CONFIGS:
            data = raw[cfg].get((b, v, u), [])
            vals = [iter_value(data, i) for i in range(1, 11)]
            present = [x for x in vals if x is not None]
            m = statistics.median(present) if present else None
            mad = (100*statistics.median([abs(x-m) for x in present])/m) if (present and m) else None
            row = [label if cfg=='kvmoff' else '', u if cfg=='kvmoff' else '', cfg] + vals + [m, mad]
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = THIN
                if isinstance(val, (int, float)) and ci >= 4:
                    c.number_format = '0.0000'
            ri += 1
        # blank separator row
        ri += 0
    freeze_and_filter(ws, 'D2', len(headers), ri-1)
    auto_width(ws, {
        1: 36, 2: 8, 3: 8,
        4: 11, 5: 11, 6: 11, 7: 11, 8: 11, 9: 11, 10: 11, 11: 11, 12: 11, 13: 11,
        14: 12, 15: 10,
    })

# ============ Sheet 4: README ============
def build_readme(wb):
    ws = wb.create_sheet('README', 0)
    notes = [
        ('lmbench 4-mode N=10 干净对照数据 — Phytium FTC862 + KylinOS 6.6.0-73-generic', None),
        ('', None),
        ('生成时间', '2026-06-04'),
        ('数据来源', 'results/n90-day3/n90-{kvmoff,nvhe,vhe,pkvm}-noLSM-full-cpu0.csv'),
        ('每个配置 iter 数', 'N=10'),
        ('CPU 频率', '锁 1900 MHz, governor=performance'),
        ('LSM 栈', 'capability,kycp（无 ksaf/bpf/audit）'),
        ('THP', 'never'),
        ('ASLR', '0'),
        ('绑核', 'cpu0 (taskset -c 0)'),
        ('lmbench config', 'CONFIG.host（全套 BENCHMARK_*=YES）'),
        ('噪声抑制', '72 项系统 + 12 项用户 systemd 服务 mask，桌面 multi-user.target'),
        ('', None),
        ('Sheet 说明：', None),
        ('  Highlights', '论文表 (Sci China Inf Sci) 30+ 项 + lat_mmap/mem hierarchy/bw 重点'),
        ('  All metrics', '全部 684 项 (bench, variant) 中位 + MAD% + 5 个跨配置 Δ%'),
        ('  Per-iter raw', '论文重点项的 10 个 iter 单值（用于自己重新算 stat）'),
        ('', None),
        ('Δ% 颜色编码（方向 × 强度）：', None),
        ('', None),
        ('  方向（按指标单位判断）：', None),
        ('    单位 us / ns（延迟）', '负 Δ% → 更快 → 绿；正 Δ% → 更慢 → 红'),
        ('    单位 MB/s（带宽）', '正 Δ% → 更高 → 绿；负 Δ% → 更低 → 红'),
        ('    单位 pages（TLB）', '正 Δ% → 更大 → 绿；负 Δ% → 更小 → 红'),
        ('', None),
        ('  强度：', None),
        ('    |Δ| > 10%', '深色 + 粗体（显著差异）'),
        ('    5% < |Δ| ≤ 10%', '中等色'),
        ('    2% < |Δ| ≤ 5%', '浅色'),
        ('    |Δ| ≤ 2%', '黑色（在 MAD 噪声范围内，不算差异）'),
        ('', None),
        ('关键 finding：', None),
        ('  1. host syscall fastpath 4 模式无差（≤ 1%）', None),
        ('  2. CPU 算术 4 模式绝对相等（sanity ✓）', None),
        ('  3. pkvm 在 lat_mmap 大段比其它 3 模式慢 +15-42%（stage-2 表维护）', None),
        ('  4. host EL2 (vhe) vs host EL1 (kvmoff/nvhe) 无可观测性能差', None),
        ('  5. TLB / cache 层级 / bw_mem 4 模式相等', None),
    ]
    ws['A1'] = notes[0][0]
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws.merge_cells('A1:D1')
    for i, (k, val) in enumerate(notes[1:], 2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=k and not k.startswith(' '))
        if val is not None: ws.cell(row=i, column=2, value=val)
    auto_width(ws, {1: 30, 2: 70})

# ============ build all ============
wb = Workbook()
wb.remove(wb.active)  # remove default sheet
build_readme(wb)
build_highlights(wb)
build_all_metrics(wb)
build_per_iter(wb)

out = (f'docs/n90-kvm-host/lmbench-N10-4config-{STAT_MODE}.xlsx'
       if STAT_MODE == 'mean'
       else 'docs/n90-kvm-host/lmbench-N10-4config.xlsx')
wb.save(out)
print(f'wrote {out}')
print(f'sheets: {wb.sheetnames}')

# ===== chain：用 ns 精度复测数据覆盖 lat_mmap 那几行 =====
import subprocess, sys, os
print('chaining → update-xlsx-precise-mmap.py (lat_mmap ns 精度覆盖) ...')
r = subprocess.run([sys.executable, 'scripts/update-xlsx-precise-mmap.py'],
                   cwd=os.getcwd())
if r.returncode != 0:
    print('WARN: update-xlsx-precise-mmap.py failed; '
          'xlsx 里 lat_mmap 行仍是 lmbench 整数 µs round 后的值')
