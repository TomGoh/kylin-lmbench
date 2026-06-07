#!/usr/bin/env python3
"""Build cross-machine VM ratio tables from lmbench raw txt/csv data.

The main sheet is intentionally a wide table:
  bench, variant, unit, direction,
  <absolute median columns for each environment>,
  <same-machine protected/confidential VM over regular VM ratios>.

Input can be either:
  * a result directory containing *iter*.txt files, parsed via parse-lmbench.py
  * an already parsed lmbench CSV with columns:
    env, core, iter, bench, variant, value, unit
"""
from __future__ import annotations

import csv
import ast
import statistics as st
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO = Path(__file__).resolve().parent.parent
OUT_CSV = REPO / "docs" / "vm-ratio-comparison.csv"
OUT_HIGHLIGHTS_CSV = REPO / "docs" / "vm-ratio-comparison-highlights.csv"
OUT_XLSX = REPO / "docs" / "vm-ratio-comparison.xlsx"


@dataclass(frozen=True)
class Dataset:
    key: str
    title: str
    source: Path
    machine: str
    vm_type: str
    drop_iters: set[int] = field(default_factory=set)


DATASETS = [
    Dataset(
        "n90_vhe_vm",
        "N90 VHE VM",
        REPO / "results" / "xkernel-n90-kvm-vhe-20260605-201819",
        "N90",
        "regular",
        {1},
    ),
    Dataset(
        "n90_nvhe_vm",
        "N90 NVHE VM",
        REPO / "results" / "xkernel-n90-kvm-nvhe-20260606-133103",
        "N90",
        "regular",
        {1},
    ),
    Dataset(
        "n90_pkvm_np_vm",
        "N90 pKVM NP VM",
        REPO / "results" / "xkernel-n90-pkvm-np-clean-20260606-164648",
        "N90",
        "pKVM non-protected",
        {1},
    ),
    Dataset(
        "n90_pkvm_pvm",
        "N90 pKVM pVM",
        REPO / "results" / "xkernel-n90-pkvm-pvm-20260605-163040",
        "N90",
        "pKVM protected",
        {1},
    ),
    Dataset(
        "kunpeng_regular_vm",
        "Kunpeng regular VM",
        REPO / "results" / "virtcca-regular-results",
        "Kunpeng",
        "regular",
    ),
    Dataset(
        "kunpeng_virtcca_cvm",
        "Kunpeng VirtCCA CVM",
        REPO / "results" / "virtcca-cvm-results",
        "Kunpeng",
        "confidential",
    ),
    Dataset(
        "hygon_csv_cvm",
        "Hygon CSV CVM",
        REPO / "results" / "hygon_csv_results",
        "Hygon",
        "confidential",
    ),
]


RATIOS = [
    ("n90_pkvm_pvm_over_nvhe_vm", "N90 pKVM pVM / NVHE VM", "n90_pkvm_pvm", "n90_nvhe_vm"),
    ("n90_pkvm_pvm_over_vhe_vm", "N90 pKVM pVM / VHE VM", "n90_pkvm_pvm", "n90_vhe_vm"),
    ("n90_pkvm_np_over_nvhe_vm", "N90 pKVM NP VM / NVHE VM", "n90_pkvm_np_vm", "n90_nvhe_vm"),
    (
        "kunpeng_virtcca_over_regular_vm",
        "Kunpeng VirtCCA CVM / regular VM",
        "kunpeng_virtcca_cvm",
        "kunpeng_regular_vm",
    ),
]


LOWER_IS_BETTER_UNITS = {"us", "ns", "seconds", "sec"}
HIGHER_IS_BETTER_UNITS = {"MB/s", "KB/s", "pages", "parallel", "ops/s"}

HEAD_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEAD_FONT = Font(color="FFFFFF", bold=True)
GROUP_FILL = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
GROUP_FONT = Font(bold=True)
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

SUMMARY_CATEGORIES = [
    "Processor, Processes",
    "Integer operations",
    "uint64 operations",
    "Float operations",
    "Double operations",
    "Context switching",
    "Local communication latencies",
    "Network/RPC latencies",
    "File & VM system latencies",
    "Local communication bandwidths",
    "File and mmap bandwidths",
    "Memory bandwidths",
    "Memory latencies",
    "Memory parallelism",
    "TLB",
    "Other",
]
CATEGORY_ORDER = {name: i + 1 for i, name in enumerate(SUMMARY_CATEGORIES)}


def load_reference_paper_table():
    """Reuse the exact Highlights grouping from scripts/build-xlsx.py."""
    src = (REPO / "scripts" / "build-xlsx.py").read_text()
    mod = ast.parse(src)
    for node in mod.body:
        if isinstance(node, ast.Assign):
            for target in node.targets:
                if isinstance(target, ast.Name) and target.id == "PAPER_TABLE":
                    return ast.literal_eval(node.value)
    raise RuntimeError("PAPER_TABLE not found in scripts/build-xlsx.py")


PAPER_TABLE = load_reference_paper_table()
PAPER_CATEGORY_ORDER = {section: i + 1 for i, (section, _items) in enumerate(PAPER_TABLE)}


def direction(unit: str) -> str:
    if unit in LOWER_IS_BETTER_UNITS:
        return "lower-is-better"
    if unit in HIGHER_IS_BETTER_UNITS:
        return "higher-is-better"
    return "unknown"


def is_lower_better(unit: str) -> bool:
    return unit in LOWER_IS_BETTER_UNITS


def delta_font(delta_pct: float | None, unit: str) -> Font:
    if delta_pct is None:
        return Font(color="000000")
    mag = abs(delta_pct)
    if mag <= 2:
        return Font(color="000000")
    is_good = delta_pct < 0 if is_lower_better(unit) else delta_pct > 0
    if mag > 10:
        return Font(color="006100" if is_good else "9C0006", bold=True)
    if mag > 5:
        return Font(color="2E7D32" if is_good else "C00000")
    return Font(color="66BB6A" if is_good else "E26B0A")


def section_order(category: str) -> str:
    if category in PAPER_CATEGORY_ORDER:
        return str(PAPER_CATEGORY_ORDER[category])
    return str(CATEGORY_ORDER.get(category, ""))


def category_for(bench: str, variant: str) -> str:
    """Classify metrics in the same broad order as lmbench getsummary."""
    if bench in {"lat_syscall", "lat_sig", "lat_proc", "lat_select"}:
        return "Processor, Processes"
    if bench == "lat_ops":
        if variant.startswith("integer_"):
            return "Integer operations"
        if variant.startswith(("int64_", "uint64_")):
            return "uint64 operations"
        if variant.startswith("float_"):
            return "Float operations"
        if variant.startswith("double_"):
            return "Double operations"
        return "Processor, Processes"
    if bench == "par_ops":
        if variant.startswith("integer_"):
            return "Integer operations"
        if variant.startswith(("int64_", "uint64_")):
            return "uint64 operations"
        if variant.startswith("float_"):
            return "Float operations"
        if variant.startswith("double_"):
            return "Double operations"
        return "Processor, Processes"
    if bench == "lat_ctx":
        return "Context switching"
    if bench in {"lat_pipe", "lat_unix"}:
        return "Local communication latencies"
    if bench in {"lat_udp", "lat_tcp", "lat_connect", "lat_rpc"}:
        return "Network/RPC latencies"
    if bench in {"lat_fs", "lat_mmap", "lat_pagefault"}:
        return "File & VM system latencies"
    if bench in {"bw_pipe", "bw_unix", "bw_tcp"}:
        return "Local communication bandwidths"
    if bench in {"lmdd", "bw_file_wr", "bw_file_rd", "bw_file_rd_o2c", "bw_mmap_rd", "bw_mmap_rd_o2c"}:
        return "File and mmap bandwidths"
    if bench.startswith("bw_mem_") or bench in {"stream", "stream2"}:
        return "Memory bandwidths"
    if bench.startswith("lat_mem_rd_"):
        return "Memory latencies"
    if bench == "par_mem":
        return "Memory parallelism"
    if bench == "tlb":
        return "TLB"
    return "Other"


def summary_order(bench: str, variant: str) -> int:
    return CATEGORY_ORDER[category_for(bench, variant)]


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as f:
        return list(csv.DictReader(f))


def parse_txt_dir(path: Path) -> list[dict[str, str]]:
    files = sorted(path.glob("*iter*.txt"))
    if not files:
        raise FileNotFoundError(f"no *iter*.txt files under {path}")
    cmd = [sys.executable, str(REPO / "parse-lmbench.py"), *map(str, files)]
    proc = subprocess.run(cmd, cwd=REPO, text=True, capture_output=True, check=True)
    return list(csv.DictReader(proc.stdout.splitlines()))


def load_dataset(ds: Dataset) -> dict[tuple[str, str, str], list[float]]:
    rows = load_csv(ds.source) if ds.source.is_file() else parse_txt_dir(ds.source)
    grouped: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    for row in rows:
        try:
            iter_id = int(row["iter"])
            if iter_id in ds.drop_iters:
                continue
            key = (row["bench"], row["variant"], row["unit"])
            grouped[key].append(float(row["value"]))
        except (KeyError, TypeError, ValueError):
            continue
    return grouped


def median(values: list[float]) -> float | None:
    return st.median(values) if values else None


def mad_pct(values: list[float]) -> float | None:
    if not values:
        return None
    m = st.median(values)
    if m == 0:
        return None
    return 100.0 * st.median(abs(x - m) for x in values) / m


def pct(num: float | None, den: float | None) -> float | None:
    if num is None or den in (None, 0):
        return None
    return 100.0 * (num - den) / den


def ratio(num: float | None, den: float | None) -> float | None:
    if num is None or den in (None, 0):
        return None
    return num / den


def fmt(value: float | None) -> str:
    return "" if value is None else f"{value:.6f}"


def build_medians(all_data: dict[str, dict[tuple[str, str, str], list[float]]]):
    return {
        ds.key: {key: median(vals) for key, vals in all_data[ds.key].items()}
        for ds in DATASETS
    }


def build_row(key, medians, label: str | None = None, category: str | None = None):
    bench, variant, unit = key
    category = category or category_for(bench, variant)
    row: dict[str, str] = {
        "summary_order": section_order(category),
        "category": category,
    }
    if label is not None:
        row["item"] = label
    row.update(
        {
            "bench": bench,
            "variant": variant,
            "unit": unit,
            "direction": direction(unit),
        }
    )
    for ds in DATASETS:
        row[f"{ds.key}_median"] = fmt(medians[ds.key].get(key))
    for ratio_key, _title, num_key, den_key in RATIOS:
        r = ratio(medians[num_key].get(key), medians[den_key].get(key))
        d = pct(medians[num_key].get(key), medians[den_key].get(key))
        row[f"{ratio_key}_ratio"] = fmt(r)
        row[f"{ratio_key}_delta_pct"] = fmt(d)
    return row


def build_all_rows(all_data: dict[str, dict[tuple[str, str, str], list[float]]], medians):
    all_keys = sorted(
        set().union(*(set(data) for data in all_data.values())),
        key=lambda k: (summary_order(k[0], k[1]), k[0], k[1], k[2]),
    )
    return [build_row(key, medians) for key in all_keys]


def build_highlight_rows(medians):
    rows = []
    for category, items in PAPER_TABLE:
        for label, bench, variant, unit in items:
            rows.append(build_row((bench, variant, unit), medians, label=label, category=category))
    return rows


def build_stats_rows(all_data: dict[str, dict[tuple[str, str, str], list[float]]]):
    rows = []
    for ds in DATASETS:
        for key in sorted(
            all_data[ds.key],
            key=lambda k: (summary_order(k[0], k[1]), k[0], k[1], k[2]),
        ):
            vals = all_data[ds.key][key]
            bench, variant, unit = key
            category = category_for(bench, variant)
            rows.append(
                {
                    "dataset": ds.key,
                    "title": ds.title,
                    "machine": ds.machine,
                    "vm_type": ds.vm_type,
                    "summary_order": CATEGORY_ORDER[category],
                    "category": category,
                    "bench": bench,
                    "variant": variant,
                    "unit": unit,
                    "n": len(vals),
                    "median": median(vals),
                    "MAD_pct": mad_pct(vals),
                }
            )
    return rows


def build_dataset_rows():
    rows = []
    for ds in DATASETS:
        if ds.source.is_file():
            source_kind = "csv"
            files = str(ds.source.relative_to(REPO))
        else:
            source_kind = "txt-dir"
            files = f"{ds.source.relative_to(REPO)}/*iter*.txt"
        rows.append(
            {
                "dataset": ds.key,
                "title": ds.title,
                "machine": ds.machine,
                "vm_type": ds.vm_type,
                "source_kind": source_kind,
                "source": files,
                "drop_iters": ",".join(map(str, sorted(ds.drop_iters))) if ds.drop_iters else "",
            }
        )
    return rows


def write_csv(rows: list[dict[str, str]], path: Path) -> None:
    headers = list(rows[0])
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def wrap_sheet(ws, header_height: int = 36, row_height: int = 24) -> None:
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = header_height if row[0].row == 1 else row_height
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def merge_repeated_cells(ws, col: int, start_row: int, end_row: int) -> None:
    if end_row <= start_row:
        return
    run_start = start_row
    previous = ws.cell(start_row, col).value
    for row in range(start_row + 1, end_row + 2):
        current = ws.cell(row, col).value if row <= end_row else None
        if current != previous:
            if previous not in (None, "") and row - run_start > 1:
                ws.merge_cells(start_row=run_start, start_column=col, end_row=row - 1, end_column=col)
                ws.cell(run_start, col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            run_start = row
            previous = current


def append_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wrap_sheet(ws)
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, 1):
        width = max(12, min(34, len(str(header)) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def append_grouped_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    group_fill = PatternFill("solid", fgColor="D9E1F2")
    current_category = None
    for row in rows:
        category = row.get("category")
        if category != current_category:
            current_category = category
            ws.append([f"{row.get('summary_order', '')}. {category}"] + [""] * (len(headers) - 1))
            group_row = ws.max_row
            ws.merge_cells(start_row=group_row, start_column=1, end_row=group_row, end_column=len(headers))
            cell = ws.cell(group_row, 1)
            cell.font = Font(bold=True)
            cell.fill = group_fill
        ws.append([row.get(h, "") for h in headers])

    wrap_sheet(ws)
    ws.freeze_panes = "G2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for idx, header in enumerate(headers, 1):
        width = max(12, min(34, len(str(header)) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def cell_number(value):
    if value in ("", None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def write_highlights_sheet(ws, rows):
    headers = (
        ["维度", "具体测试项", "unit", "direction"]
        + [ds.title for ds in DATASETS]
        + [col for _key, title, _num, _den in RATIOS for col in (title, f"Δ% {title}")]
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[1].height = 32

    row_by_section: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        row_by_section[row["category"]].append(row)

    excel_row = 2
    for section, _items in PAPER_TABLE:
        c = ws.cell(row=excel_row, column=1, value=section)
        c.fill = GROUP_FILL
        c.font = GROUP_FONT
        for ci in range(2, len(headers) + 1):
            ws.cell(row=excel_row, column=ci).fill = GROUP_FILL
        excel_row += 1

        section_start = excel_row
        for row in row_by_section.get(section, []):
            values = ["", row["item"], row["unit"], row["direction"]]
            values.extend(cell_number(row[f"{ds.key}_median"]) for ds in DATASETS)
            for ratio_key, _title, _num_key, _den_key in RATIOS:
                values.append(cell_number(row[f"{ratio_key}_ratio"]))
                values.append(cell_number(row[f"{ratio_key}_delta_pct"]))

            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=ci, value=val)
                cell.border = THIN
                if isinstance(val, (int, float)):
                    cell.number_format = "+0.00;-0.00" if ci > 4 + len(DATASETS) else "0.0000"
                    if ci > 4 + len(DATASETS) and (ci - 4 - len(DATASETS)) % 2 == 0:
                        cell.font = delta_font(float(val), row["unit"])
            excel_row += 1
        merge_repeated_cells(ws, col=4, start_row=section_start, end_row=excel_row - 1)

    wrap_sheet(ws)
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{excel_row - 1}"
    widths = {1: 24, 2: 36, 3: 8, 4: 14}
    for ci in range(5, 5 + len(DATASETS)):
        widths[ci] = 16
    for ci in range(5 + len(DATASETS), len(headers) + 1):
        widths[ci] = 22
    for ci, width in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = width


def write_all_metrics_sheet(ws, rows):
    headers = (
        ["bench", "variant", "unit", "direction"]
        + [f"{ds.title} median" for ds in DATASETS]
        + [col for _key, title, _num, _den in RATIOS for col in (f"{title} ratio", f"Δ% {title}")]
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[1].height = 32

    for row in rows:
        values = [row["bench"], row["variant"], row["unit"], row["direction"]]
        values.extend(cell_number(row[f"{ds.key}_median"]) for ds in DATASETS)
        for ratio_key, _title, _num_key, _den_key in RATIOS:
            values.append(cell_number(row[f"{ratio_key}_ratio"]))
            values.append(cell_number(row[f"{ratio_key}_delta_pct"]))
        ws.append(values)
        excel_row = ws.max_row
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=excel_row, column=ci)
            cell.border = THIN
            if isinstance(cell.value, (int, float)):
                cell.number_format = "+0.00;-0.00" if ci > 4 + len(DATASETS) else "0.0000"
                if ci > 4 + len(DATASETS) and (ci - 4 - len(DATASETS)) % 2 == 0:
                    cell.font = delta_font(float(cell.value), row["unit"])

    wrap_sheet(ws)
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    widths = {1: 22, 2: 28, 3: 8, 4: 14}
    for ci in range(5, 5 + len(DATASETS)):
        widths[ci] = 16
    for ci in range(5 + len(DATASETS), len(headers) + 1):
        widths[ci] = 22
    for ci, width in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = width


def write_xlsx(highlight_rows, all_rows, stats_rows, dataset_rows, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Highlights"
    write_highlights_sheet(ws, highlight_rows)

    ws = wb.create_sheet("All metrics")
    write_all_metrics_sheet(ws, all_rows)

    ws = wb.create_sheet("Raw stats")
    headers = [
        "dataset",
        "title",
        "machine",
        "vm_type",
        "summary_order",
        "category",
        "bench",
        "variant",
        "unit",
        "n",
        "median",
        "MAD_pct",
    ]
    append_grouped_table(ws, headers, stats_rows)

    ws = wb.create_sheet("Datasets")
    headers = ["dataset", "title", "machine", "vm_type", "source_kind", "source", "drop_iters"]
    append_table(ws, headers, dataset_rows)

    ws = wb.create_sheet("README")
    notes = [
        ("输出文件", str(path.relative_to(REPO))),
        ("全量 CSV", str(OUT_CSV.relative_to(REPO))),
        ("重点项 CSV", str(OUT_HIGHLIGHTS_CSV.relative_to(REPO))),
        ("统计口径", "median；Raw stats sheet 同时给出 MAD%"),
        ("N90 口径", "x-kernel guest 数据丢弃 iter 1，用 iter 2-10 计算稳定态中位数"),
        ("Kunpeng 口径", "regular VM 与 VirtCCA CVM 都使用 iter 1-10"),
        ("Hygon 口径", "当前只有 CSV CVM 绝对值；普通 VM 基线未纳入，因此没有 Hygon ratio"),
        ("ratio 含义", "受保护或机密 VM median / 同机普通 VM median"),
        ("delta_pct 含义", "(受保护或机密 VM median - 同机普通 VM median) / 同机普通 VM median * 100"),
        ("Highlights", "按 lmbench-N10-4config.xlsx 的重点项规模组织，当前 45 个指标"),
        ("All metrics", "保留全部可解析指标，按 lmbench scripts/getsummary 的大类顺序排列"),
    ]
    ws.append(["item", "value"])
    for item, value in notes:
        ws.append([item, value])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_sheet(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100

    wb.save(path)


def main() -> None:
    all_data = {ds.key: load_dataset(ds) for ds in DATASETS}
    medians = build_medians(all_data)
    all_rows = build_all_rows(all_data, medians)
    highlight_rows = build_highlight_rows(medians)
    stats_rows = build_stats_rows(all_data)
    dataset_rows = build_dataset_rows()
    write_csv(all_rows, OUT_CSV)
    write_csv(highlight_rows, OUT_HIGHLIGHTS_CSV)
    write_xlsx(highlight_rows, all_rows, stats_rows, dataset_rows, OUT_XLSX)
    print(f"wrote {OUT_CSV.relative_to(REPO)}")
    print(f"wrote {OUT_HIGHLIGHTS_CSV.relative_to(REPO)}")
    print(f"wrote {OUT_XLSX.relative_to(REPO)}")
    for ds in DATASETS:
        total_samples = sum(len(v) for v in all_data[ds.key].values())
        print(f"{ds.key}: metrics={len(all_data[ds.key])} samples={total_samples}")


if __name__ == "__main__":
    main()
