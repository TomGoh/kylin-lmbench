#!/usr/bin/env python3
"""把 precise lat_mmap (ns 精度) 数据写入两份 xlsx 表格，替换 N=10 整数 µs 行。

更新位置（两份 xlsx 同样的 sheet/row 布局）：
  Sheet "Highlights"     rows 41-44   (0.5/1/16/67 MB)
  Sheet "All metrics"    rows 585-592 (8 个 lat_mmap size 行，包含 33MB 我们没动)
  Sheet "Per-iter raw"   rows 130-145 (0.5/1/16/67 MB × 4 mode × 10 iter)

输入数据：results/precise-mmap/<mode>.log
"""
from __future__ import annotations
import re
import statistics as st
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
LOG_DIR = REPO / "results" / "precise-mmap"
MODES = ["kvmoff", "vhe", "nvhe", "pkvm"]
LINE_RE = re.compile(
    r"size_mb=([\d.]+) iters=\d+ total_ns=\d+ per_iter_ns=[\d.]+ per_iter_us=([\d.]+)"
)


def parse_log(path: Path) -> dict[float, list[float]]:
    by_size: dict[float, list[float]] = {}
    for line in path.read_text().splitlines():
        m = LINE_RE.search(line)
        if not m:
            continue
        sz = float(m.group(1))
        us = float(m.group(2))
        by_size.setdefault(sz, []).append(us)
    return by_size


def mad_pct(xs):
    if not xs:
        return 0.0
    m = st.median(xs)
    if m == 0:
        return 0.0
    return 100.0 * st.median(abs(x - m) for x in xs) / m


def stddev_pct(xs):
    if len(xs) < 2:
        return 0.0
    m = st.mean(xs)
    if m == 0:
        return 0.0
    return 100.0 * st.stdev(xs) / m


def agg(xs, kind):
    if not xs:
        return float("nan"), float("nan")
    if kind == "median":
        return st.median(xs), mad_pct(xs)
    return st.mean(xs), stddev_pct(xs)


def pct(a, b):
    return 100.0 * (a - b) / b if b else 0.0


def update(xlsx_path: Path, kind: str, data: dict):
    wb = load_workbook(xlsx_path)

    # ===== Highlights =====
    # cols: A=维度 B=项 C=unit D=kvmoff E=nvhe F=vhe G=pkvm
    #       H=Δnvhe/kvmoff I=Δvhe/kvmoff J=Δpkvm/kvmoff K=Δpkvm/vhe L=Δpkvm/nvhe
    ws = wb["Highlights"]
    HI = {41: 0.5, 42: 1, 43: 16, 44: 64}  # row -> size_mb
    for row, sz in HI.items():
        vals = {m: agg(data[m][sz], kind)[0] for m in MODES}
        ws.cell(row=row, column=4).value = vals["kvmoff"]
        ws.cell(row=row, column=5).value = vals["nvhe"]
        ws.cell(row=row, column=6).value = vals["vhe"]
        ws.cell(row=row, column=7).value = vals["pkvm"]
        ws.cell(row=row, column=8).value = pct(vals["nvhe"], vals["kvmoff"])
        ws.cell(row=row, column=9).value = pct(vals["vhe"], vals["kvmoff"])
        ws.cell(row=row, column=10).value = pct(vals["pkvm"], vals["kvmoff"])
        ws.cell(row=row, column=11).value = pct(vals["pkvm"], vals["vhe"])
        ws.cell(row=row, column=12).value = pct(vals["pkvm"], vals["nvhe"])

    # ===== All metrics =====
    # header: A=bench B=variant C=unit
    #         D=kvmoff_med E=kvmoff_var% F=nvhe_med G=nvhe_var%
    #         H=vhe_med I=vhe_var% J=pkvm_med K=pkvm_var%
    #         L=Δ nvhe/kvmoff M=Δ vhe/kvmoff N=Δ pkvm/kvmoff
    #         O=Δ pkvm/vhe P=Δ pkvm/nvhe
    ws = wb["All metrics"]
    AM = {
        585: 0.5,
        586: 1,
        587: 16,
        588: 2,
        # 589 is 33.554432 MB (32 MB) — we DON'T have precise data, skip
        590: 4,
        591: 64,  # original label sz67.108864MB == 64 MB precise
        592: 8,
    }
    for row, sz in AM.items():
        ag = {m: agg(data[m][sz], kind) for m in MODES}
        ws.cell(row=row, column=4).value = ag["kvmoff"][0]
        ws.cell(row=row, column=5).value = ag["kvmoff"][1]
        ws.cell(row=row, column=6).value = ag["nvhe"][0]
        ws.cell(row=row, column=7).value = ag["nvhe"][1]
        ws.cell(row=row, column=8).value = ag["vhe"][0]
        ws.cell(row=row, column=9).value = ag["vhe"][1]
        ws.cell(row=row, column=10).value = ag["pkvm"][0]
        ws.cell(row=row, column=11).value = ag["pkvm"][1]
        ws.cell(row=row, column=12).value = pct(ag["nvhe"][0], ag["kvmoff"][0])
        ws.cell(row=row, column=13).value = pct(ag["vhe"][0], ag["kvmoff"][0])
        ws.cell(row=row, column=14).value = pct(ag["pkvm"][0], ag["kvmoff"][0])
        ws.cell(row=row, column=15).value = pct(ag["pkvm"][0], ag["vhe"][0])
        ws.cell(row=row, column=16).value = pct(ag["pkvm"][0], ag["nvhe"][0])

    # ===== Per-iter raw =====
    # A=label B=unit C=mode D-M=10 iter values N=agg O=var%
    ws = wb["Per-iter raw"]
    PI = {
        130: (0.5, "kvmoff"), 131: (0.5, "nvhe"), 132: (0.5, "vhe"), 133: (0.5, "pkvm"),
        134: (1,   "kvmoff"), 135: (1,   "nvhe"), 136: (1,   "vhe"), 137: (1,   "pkvm"),
        138: (16,  "kvmoff"), 139: (16,  "nvhe"), 140: (16,  "vhe"), 141: (16,  "pkvm"),
        142: (64,  "kvmoff"), 143: (64,  "nvhe"), 144: (64,  "vhe"), 145: (64,  "pkvm"),
    }
    for row, (sz, mode) in PI.items():
        xs = data[mode][sz]
        assert len(xs) == 10, f"expected 10 iters for {mode} {sz}MB, got {len(xs)}"
        for i, v in enumerate(xs):
            ws.cell(row=row, column=4 + i).value = v
        ag_val, var = agg(xs, kind)
        ws.cell(row=row, column=14).value = ag_val
        ws.cell(row=row, column=15).value = var

    wb.save(xlsx_path)
    print(f"Updated {xlsx_path.name} ({kind})")


def main():
    data: dict[str, dict[float, list[float]]] = {}
    for mode in MODES:
        data[mode] = parse_log(LOG_DIR / f"{mode}.log")
    for mode in MODES:
        for sz in [0.5, 1, 2, 4, 8, 16, 64]:
            n = len(data[mode].get(sz, []))
            assert n == 10, f"{mode} {sz}MB has {n} iters, expected 10"

    update(REPO / "docs" / "findings-2026-06-03" / "lmbench-N10-4config.xlsx", "median", data)
    update(REPO / "docs" / "findings-2026-06-03" / "lmbench-N10-4config-mean.xlsx", "mean", data)


if __name__ == "__main__":
    main()
